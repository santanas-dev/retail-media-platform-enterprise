#!/usr/bin/env python3
"""Tamper tests for ROADMAP-GUARD-002.

Creates temporary copies of roadmap.xlsx with deliberate violations,
runs the guard, and asserts violations are detected.

Tests:
  1. UNDERSTATE: change G1 story from ✅ → ⚪️ → detected
  2. OVERCLAIM: set Итог=✅ on blocked row → detected
  3. CLEAN: unmodified → 0 findings
"""

import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
ROADMAP_ORIG = REPO_ROOT / "docs" / "product" / "roadmap-s020-2026-07-10.xlsx"
GUARD_SCRIPT = REPO_ROOT / "scripts" / "roadmap-consistency-check.py"

passed = 0
failed = 0


def run_guard_on(tamper_path):
    """Copy tampered file over real, run guard, restore."""
    backup = str(ROADMAP_ORIG) + ".bak"
    shutil.copy2(ROADMAP_ORIG, backup)
    try:
        shutil.copy2(tamper_path, ROADMAP_ORIG)
        result = subprocess.run(
            [sys.executable, str(GUARD_SCRIPT)],
            capture_output=True, text=True, timeout=30,
        )
        return result.stdout
    finally:
        shutil.copy2(backup, ROADMAP_ORIG)
        os.remove(backup)


def test(name, tamper_fn, expect_violation=True):
    global passed, failed
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        shutil.copy2(ROADMAP_ORIG, tmp.name)
        tp = tmp.name

    try:
        wb = openpyxl.load_workbook(tp)
        tamper_fn(wb)
        wb.save(tp)
        stdout = run_guard_on(tp)
        clean = "0 violations" in stdout

        if expect_violation and not clean:
            print(f"  ✅ PASS: {name}")
            passed += 1
        elif not expect_violation and clean:
            print(f"  ✅ PASS: {name} (clean)")
            passed += 1
        else:
            print(f"  ❌ FAIL: {name}")
            print(f"     Expected: {'violation' if expect_violation else 'clean'}")
            print(f"     Got: {'clean' if clean else 'violations found'}")
            failed += 1
    finally:
        try:
            os.unlink(tp)
        except OSError:
            pass


ws_name = "Бизнес-функции Roadmap"

print("=== ROADMAP-GUARD-002 Tamper Tests ===\n")


# Test 1: Understate G1
def understate_g1(wb):
    ws = wb[ws_name]
    for row in range(2, ws.max_row + 1):
        func = str(ws.cell(row=row, column=2).value or "")
        if "Создание и редактирование" in func:
            old = str(ws.cell(row=row, column=5).value or "")
            ws.cell(row=row, column=5).value = old.replace(
                "✅ campaign.create", "⚪️ campaign.create"
            )
            return

test("Understate G1 (campaign.create ✅→⚪️)", understate_g1)


# Test 2: Overclaim blocked row
def overclaim_blocked(wb):
    ws = wb[ws_name]
    for row in range(2, ws.max_row + 1):
        func = str(ws.cell(row=row, column=2).value or "")
        if "Согласование" in func:
            ws.cell(row=row, column=6).value = "✅ Готово/Юзабельно"
            ws.cell(row=row, column=3).value = "✅"
            ws.cell(row=row, column=4).value = "✅"
            ws.cell(row=row, column=5).value = (
                "✅ campaign.approve / ✅ campaign.reject"
            )
            return

test("Overclaim (approval Итог=✅ without smoke)", overclaim_blocked)


# Test 3: Clean
test("Clean workbook → 0 findings", lambda wb: None, expect_violation=False)

print(f"\n=== Results: {passed} passed, {failed} failed ===")
sys.exit(0 if failed == 0 else 1)
