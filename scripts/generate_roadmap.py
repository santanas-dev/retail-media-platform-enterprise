#!/usr/bin/env python3
"""S-020: Generate roadmap Excel with Technical and Business tabs."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ===========================================================================
# Styles
# ===========================================================================
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
section_font = Font(bold=True, size=11, color="1F3864")

done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
done_font = Font(color="006100", bold=True)
pilot_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
pilot_font = Font(color="9C6500", bold=True)
partial_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
partial_font = Font(color="974706", bold=True)
not_started_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
not_started_font = Font(color="808080")
deferred_fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
deferred_font = Font(color="5B3A8C")

thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)
wrap_align = Alignment(wrap_text=True, vertical='top')

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = thin_border

def style_section_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
        cell.alignment = wrap_align

def write_status(ws, row, col, status):
    cell = ws.cell(row=row, column=col)
    if status == "✅ Готово":
        cell.fill = done_fill; cell.font = done_font
    elif status == "🟡 Pilot-ready":
        cell.fill = pilot_fill; cell.font = pilot_font
    elif status == "🟠 Частично":
        cell.fill = partial_fill; cell.font = partial_font
    elif status == "🚫 Deferred":
        cell.fill = deferred_fill; cell.font = deferred_font
    else:
        cell.fill = not_started_fill; cell.font = not_started_font
    cell.value = status
    cell.alignment = wrap_align
    cell.border = thin_border

def write_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.alignment = wrap_align
    cell.border = thin_border

# ===========================================================================
# Sheet 1: Технический Roadmap
# ===========================================================================
ws1 = wb.active
ws1.title = "Технический Roadmap"
ws1.sheet_properties.tabColor = "2F5496"

cols1 = [
    ("A", "Модуль / Подсистема", 38),
    ("B", "Статус", 18),
    ("C", "Коммиты / Тесты / ADR", 48),
    ("D", "Следующий шаг", 36),
    ("E", "Заметки", 36),
]
for i, (cl, title, w) in enumerate(cols1, 1):
    ws1.cell(row=1, column=i, value=title)
    ws1.column_dimensions[cl].width = w
style_header_row(ws1, 1, 5)

tech = [
    ("SECTION", "Ядро платформы", "", "", ""),
    ("Импорт-границы (ADR-014)", "✅ Готово",
     "check-import-boundaries.py: 44/44.", "—", "Layer discipline enforced."),
    ("Outbox Foundation (ADR-011)", "✅ Готово",
     "Migration 007. OutboxEvent model. 10 unit + 9 behavioral.", "—", "Transactional outbox pattern."),
    ("DB Role Architecture (S-019)", "✅ Готово",
     "b05fa55. 3 роли: retail_media_owner + retail_media_app (NOBYPASSRLS). init-db.sql + grant-app-role.py.",
     "Worker per-scope resolution (Phase 3.6)", "Compose/CI parity achieved."),
    ("Release Versioning (S-015)", "✅ Готово",
     "release-versioning.md. Tag v0.1-admin-campaign-mvp (d291cfed).",
     "v0.2-media-upload-mvp tag (pending approval)", "Release policy defined."),

    ("SECTION", "Безопасность и доступ", "", "", ""),
    ("Auth / Dual Login (S-016)", "✅ Готово",
     "local_advertiser + local_break_glass (bcrypt). AD stub → honest 503. /me DB-backed. Seed production gate.",
     "Real LDAPS wiring (not in pilot)", "3 auth_provider types: ad, local_advertiser, local_break_glass."),
    ("RBAC / RLS (S-008, ADR-009)", "✅ Готово",
     "28 RLS policies (7 tables × 4). CI: retail_media_app NOBYPASSRLS gate. 4 behavioral RLS enforcement tests.",
     "Worker per-scope resolution (ADR-009 §9)", "Two-layer defense: app RBAC + DB RLS."),
    ("CORS Baseline (S-011)", "✅ Готово",
     "SecurityConfig. Dev: localhost:5173. Production: explicit origins required. 10 tests.",
     "—", "control-api only. device-gateway CORS removed in S-019."),
    ("Rate Limiting (S-010)", "✅ Готово",
     "5 attempts / 15 min window. 429 response. No user enumeration. 5 unit + 4 behavioral.",
     "—", "Pre-user-lookup guard (no timing side-channel)."),

    ("SECTION", "Campaign Domain", "", "", ""),
    ("Campaign CRUD (S-003, S-004)", "✅ Готово",
     "7 endpoints, 7 ORM models, migration 006. 41 unit + 30 behavioral.",
     "—", "Tenant isolation, cross-org validation, outbox integration."),
    ("Flight/Placement/Creative APIs (B1)", "✅ Готово",
     "POST/PATCH flights, placements, creatives. Flight window validation. Creative scope fix. 42 behavioral.",
     "—", "Outbox per ADR-015."),
    ("Approval Workflow (S-005)", "✅ Готово",
     "request-approval/approve/reject. Status transitions, approval records, audit trail. 18 unit + 24 behavioral.",
     "—", "Idempotent. requested_at semantics. Contract validation."),
    ("Creative Media Upload (S-017)", "✅ Готово",
     "MinIO presigned URL flow: upload-intent → PUT → complete-upload. Server-side SHA-256. Admin-web upload UI with progress bar. 818 unit + admin-web UI shipped.",
     "Malware scan, transcoding, CDN (deferred)", "Pilot: CREATIVE_AUTO_APPROVE_UPLOADS=true. Approval gate requires real upload."),
    ("PoP Ingestion (S-007c)", "✅ Готово",
     "POST /api/v1/pop/batch. Device JWT, 11-step validation, quarantine 72h, billing-grade accepted. 29 unit + 27 behavioral.",
     "ClickHouse materialized views (4.3e)", "Dedup by event_id. Clock drift ±5min."),
    ("PoP Reporting (S-007d)", "✅ Готово",
     "GET pop/{summary,by-day,by-surface}. Scoped advertiser. Campaign ownership guard. 13 unit + 23 behavioral.",
     "Export PDF/XLSX/CSV (deferred)", "No PII/secrets in responses."),

    ("SECTION", "Канал доставки (Delivery Pipeline)", "", "", ""),
    ("Manifest Generation (S-006c)", "✅ Готово",
     "delivery.py (776 lines): eligibility, target resolution, manifest JSON gen. Flat playlist v1. 25 unit + 12 behavioral.",
     "Production HMAC signing (placeholder)", "compute_manifest_id + compute_campaign_version_hash."),
    ("Device Gateway (S-006d)", "✅ Готово",
     "GET /api/v1/device/manifest/latest. Device JWT auth, ETag/304, security-hardened. 10 unit + 7 behavioral.",
     "—", "No storage secrets in responses. Store-orphan detection."),
    ("NATS Delivery Pipeline (B2, S-006f)", "✅ Готово",
     "outbox → relay → NATS JetStream → consumer → manifest. E2E test. 82+6+1 tests (opt-in).",
     "Production provisioning config", "RUN_NATS_INTEGRATION_TESTS=1."),
    ("Pilot E2E Smoke (B3, S-006g)", "✅ Готово",
     "Campaign → outbox → relay → NATS → consumer → manifest → device-gateway HTTP (200+ETag+304). 1 E2E (opt-in).",
     "—", "Full chain: 17 manifest fields verified, no secrets."),
    ("Runtime Simulator (S-006e)", "✅ Готово",
     "ADR-013 headless simulator: manifest apply, kill-switch (4 levels, fail-closed), render slot (6 gates), PoP integrity, offline TTL. 41 tests.",
     "Real KSO player integration", "Proves safety invariants before real hardware."),
    ("Manifest/PoP Contracts (S-018)", "✅ Готово",
     "manifest_v1.schema.json + proof_event_v1.schema.json (flat). Simulator: playback_result + campaign_id. 15+16 contract tests.",
     "Nested per-surface playlist → Manifest v2", "Full-chain: manifest → simulator → PoP → PopBatchRequest."),

    ("SECTION", "Observability и эксплуатация", "", "", ""),
    ("Health / Readiness (S-013, S-014)", "✅ Готово",
     "4× /health endpoints. Thread-safe HealthState. Graceful shutdown SIGTERM. Dead-letter counter. 86+35 unit tests.",
     "Prometheus metrics, alert rules (deferred)", "Periodic summary logger (60s)."),
    ("DB Readiness Checks (S-014)", "✅ Готово",
     "check_db_health() at startup. check_db_role_safety(). Fail-fast with safe messages.",
     "—", "No DATABASE_URL in error output."),

    ("SECTION", "Frontend", "", "", ""),
    ("Admin Web Auth Shell (S-009b)", "🟠 Частично",
     "React 19, react-router-dom, vitest. Login/logout/session, protected routes, /me, sidebar nav. 17 tests. Build: 48 modules, 605ms.",
     "S-009c: wire campaign list/detail to real API", "Auth contract: HttpOnly cookie, no refresh in JSON."),
    ("Admin Campaign UI", "🟠 Частично",
     "Pages exist as placeholders. Create wizard placeholder. Creative upload UI complete (S-017).",
     "Wire API data: campaigns list, detail, create wizard", "Backend APIs all ready — frontend wiring pending."),
    ("Advertiser Portal", "⚪ Не начато",
     "TZ v2.5 requires advertiser self-service portal with read-only campaign/report access.",
     "After admin-web campaign UI stable (S-009c+)", "Separate React app or route group."),

    ("SECTION", "Каналы и устройства (носители)", "", "", ""),
    ("KSO Player (Linux/Chromium)", "⚪ Не начато",
     "ADR-013: edge safety architecture locked. Runtime simulator: 41 safety proofs. Device Gateway ready.",
     "No real KSO player binary/wrapper. No hardware integration.",
     "Real KSO player/sidecar — outside backend pilot scope."),
    ("Android / Android TV", "🚫 Deferred",
     "Architecture: capability profiles. Manifest supports multiple device_types.",
     "No Android player. No WebView kiosk mode.",
     "After KSO pilot (channel 2)."),
    ("Price Checker (Android)", "🚫 Deferred",
     "Architecture in TZ §6.8.",
     "No implementation.",
     "After KSO + Android TV."),
    ("ESL / Electronic Price Tags", "🚫 Deferred",
     "Architecture: gateway adapters to vendor API (TZ §6.8).",
     "No ESL gateway adapters.",
     "Separate integration project."),
    ("LED Shelf Banners", "🚫 Deferred",
     "Architecture in TZ.",
     "No implementation.",
     "Low priority."),

    ("SECTION", "Аналитика, инвентарь, отчётность", "", "", ""),
    ("ClickHouse / Materialized Views (4.3e)", "🚫 Deferred",
     "ADR-007: deferred to Phase 4+. Compose: container present, not wired.",
     "No analytical queries. No materialized views. No ClickHouse schema.",
     "After PoP pipeline stable."),
    ("Inventory / Reservations / Forecasting", "⚪ Не начато",
     "TZ §6.3: availability grid, forecasting, prime-time rules, statuses.",
     "No inventory calculation. No forecasting engine.",
     "Planning after campaign domain stable."),
    ("Export (PDF/XLSX/CSV)", "⚪ Не начато",
     "TZ requires export for reports.",
     "No export endpoints.",
     "After PoP reporting stable."),
    ("Billing", "🚫 Deferred",
     "TZ §2.2: billing is out of scope for first phase.",
     "No billing logic.",
     "Separate module after production pilot."),

    ("SECTION", "Emergency и эксплуатация", "", "", ""),
    ("Emergency Management", "⚪ Не начато",
     "ADR-013 kill-switch architecture proven by simulator (4 levels, fail-closed).",
     "No emergency API. No MFA for emergency users. No UI.",
     "After device gateway stable."),
    ("Kill-Switch (ADR-013)", "✅ Готово",
     "Runtime simulator: 4 levels (graceful/pause/immediate/emergency), fail-closed, reason required. 41 tests.",
     "Real KSO player integration", "Simulator proves correctness."),
    ("DR / Backups / HA", "⚪ Не начато",
     "TZ §17: backup, recovery, load testing (40K devices).",
     "No automated PostgreSQL backups. No MinIO mirroring. No DR plan.",
     "Before production deployment."),
    ("Production Observability", "🚫 Deferred",
     "Foundation: HealthState counters (published, failed, dead-letter, acked, nakd). Periodic summary logger.",
     "No Prometheus metrics. No alert rules. No Grafana dashboards.",
     "Prometheus + Grafana after compose → production deploy."),
]

row = 2
for item in tech:
    if item[0] == "SECTION":
        for c in range(1, 6):
            ws1.cell(row=row, column=c, value=item[c-1])
        style_section_row(ws1, row, 5)
    else:
        write_cell(ws1, row, 1, item[0])
        write_status(ws1, row, 2, item[1])
        write_cell(ws1, row, 3, item[2])
        write_cell(ws1, row, 4, item[3])
        write_cell(ws1, row, 5, item[4])
    row += 1

# ===========================================================================
# Sheet 2: Бизнес-функции Roadmap
# ===========================================================================
ws2 = wb.create_sheet("Бизнес-функции Roadmap")
ws2.sheet_properties.tabColor = "548235"

cols2 = [
    ("A", "Бизнес-функция", 38),
    ("B", "Статус", 18),
    ("C", "Что можно использовать сейчас", 48),
    ("D", "Что нельзя использовать", 48),
    ("E", "План / Ближайший шаг", 36),
]
for i, (cl, title, w) in enumerate(cols2, 1):
    ws2.cell(row=1, column=i, value=title)
    ws2.column_dimensions[cl].width = w
style_header_row(ws2, 1, 5)

biz = [
    ("SECTION", "Пользователи и доступ", "", "", ""),
    ("Вход сотрудников / рекламодателей", "🟡 Pilot-ready",
     "Тестовый вход через local_advertiser и local_break_glass (bcrypt).\nJWT access + refresh токены, HttpOnly cookie.\nLogin/logout/refresh — работают через API и admin-web.",
     "Реальный LDAPS (Active Directory) не подключён — возвращает честный 503.\nMFA не реализован.\nНет саморегистрации рекламодателей.",
     "Ждём provisioning LDAPS-контроллера. Пока: тестовые учётные записи для пилота."),
    ("Роли и права (RBAC)", "✅ Готово",
     "4 системные роли (system_admin, security_admin, operator, analyst).\n16 permissions с backend-проверкой.\nДва слоя защиты: app RBAC + PostgreSQL RLS.\nScoped-доступ: branch, cluster, store, advertiser.",
     "Нет UI для управления ролями и назначения прав (только API).",
     "—"),
    ("Личный кабинет рекламодателя", "⚪ Не начато",
     "—",
     "Фронтенд рекламодателя отсутствует полностью.\nТЗ требует: read-only доступ к своим кампаниям и отчётам.",
     "После стабилизации admin-web (S-009c+)."),

    ("SECTION", "Управление рекламными кампаниями", "", "", ""),
    ("Создание и редактирование кампаний", "🟡 Pilot-ready",
     "Backend API: полный CRUD кампаний (create, read, update, archive).\nFlights (периоды), placements (размещения), creatives (креативы) — все endpoint'ы работают.\nTenant isolation: рекламодатель видит только свои кампании.",
     "UI: только login/logout работает. Страницы campaign list/detail/create — заглушки, не подключены к API.\nНет календаря планирования, нет drag-and-drop, нет визуального редактора.",
     "S-009c: подключить campaign list/detail к реальным API."),
    ("Согласование кампаний (Approval)", "🟡 Pilot-ready",
     "Backend: request-approval → approve/reject. Полный audit trail: кто, когда, причина.\nIdempotent — повторный запрос безопасен.",
     "UI approval не реализован (нет кнопок «Отправить на согласование» / «Утвердить» / «Отклонить»).",
     "После campaign UI (S-009c)."),
    ("Загрузка креативов (медиафайлы)", "🟡 Pilot-ready",
     "Backend: presigned URL (MinIO/S3) с серверной проверкой SHA-256. Безопасность: storage_bucket/key не попадают в ответы API.\nAdmin-web: кнопка «Загрузить файл», прогресс-бар, сообщения об ошибках.",
     "Файлы загружаются только через admin-web (рекламодательский портал не готов).\nНет проверки на вирусы (malware scan).\nНет транскодирования и CDN.",
     "Malware scan, transcoding — отложены. Работает для пилота."),

    ("SECTION", "Доставка контента на устройства", "", "", ""),
    ("Формирование плейлистов (manifest)", "✅ Готово",
     "Manifest JSON с полями: device_id, store_id, playlist[], media_files[], valid_from/to, offline_ttl_hours.\nETag/304 — кэширование на устройстве.\nАвтоматическая генерация при изменении кампании.",
     "Подпись manifest — HMAC-заглушка (не для production).\nNested per-surface playlist — Manifest v2 (для multi-surface KSO).",
     "Production manifest signing (отдельная задача)."),
    ("Получение manifest устройством", "✅ Готово",
     "GET /api/v1/device/manifest/latest.\nАутентификация: device JWT (sub=device_id, auth_provider=device).\nПроверка статуса устройства (active/online).\nETag/If-None-Match для кэширования.",
     "Только HTTP (не HTTPS/mTLS).\nНет emergency-канала для срочных команд.",
     "mTLS после интеграции с реальным KSO плеером."),
    ("Proof-of-Play (подтверждение показов)", "✅ Готово",
     "POST /api/v1/pop/batch — приём PoP-событий от устройств.\n11-шаговая валидация: schema, device, dedup, duration, playback, clock drift, cross-entity.\nКарантин 72h. Billing-grade accepted события.",
     "PoP эмитится только симулятором (не реальным KSO плеером).\nНет экспорта отчётов о показах.",
     "Реальный KSO плеер — вне backend-пилота."),
    ("Отчёты по показам", "✅ Готово",
     "GET /api/v1/identity/campaigns/{id}/pop/{summary,by-day,by-surface}.\nScoped: рекламодатель видит только свои кампании.\nФильтрация: только accepted + campaign_verified + playback_result=success.",
     "Нет UI для отчётов.\nНет экспорта PDF/XLSX/CSV.\nНет дашборда сети (ТЗ §15.1).",
     "После campaign UI (S-009c+)."),

    ("SECTION", "Каналы и устройства", "", "", ""),
    ("КСО (кассы самообслуживания)", "⚪ Не начато",
     "Архитектура зафиксирована (ADR-013, ADR-016).\nСимулятор: 41 тест доказывает безопасность (manifest apply, kill-switch, render, PoP).\nDevice Gateway готов принимать запросы от устройств.",
     "Реального KSO плеера нет.\nНет интеграции с железом КСО (ServPlus Sherman-J 5.1 + СуперМаг УКМ 4).\nНет Linux wrapper/agent для управления плеером.",
     "Пилотный KSO плеер/сайдкар — главный следующий шаг после backend pilot."),
    ("Android / Android TV", "🚫 Deferred",
     "Архитектура заложена: capability profiles в БД, manifest поддерживает несколько device_types.",
     "Нет Android плеера.\nНет WebView kiosk режима.\nНет ТВ-приставок.",
     "Второй канал после KSO pilot."),
    ("Price Checker (Android)", "🚫 Deferred",
     "Архитектура в ТЗ (§6.8): показ рекламы только в idle/screen saver, без блокировки проверки цены.",
     "Нет реализации.",
     "После KSO + Android TV."),
    ("ESL / Электронные ценники", "🚫 Deferred",
     "Архитектура в ТЗ (§6.8): интеграция через шлюзы и vendor API.",
     "Нет адаптеров к ESL-шлюзам.\nНет поддержки протоколов вендоров.",
     "Отдельный проект интеграции."),
    ("LED Shelf Banners", "🚫 Deferred",
     "Архитектура в ТЗ.",
     "Нет реализации.",
     "Низкий приоритет."),

    ("SECTION", "Инвентарь и прогноз", "", "", ""),
    ("Инвентарь рекламных мест", "⚪ Не начато",
     "—",
     "Нет расчёта свободного/занятого/зарезервированного эфирного времени.\nНет визуальной сетки инвентаря по филиалам/магазинам.\nНет настройки правил (макс. нагрузка, prime-time, приоритеты).",
     "Планирование после стабилизации campaign domain."),
    ("Прогноз показов", "⚪ Не начато",
     "—",
     "Нет данных для прогноза — нужен работающий PoP pipeline + статистика за период.\nНет ML/статистической модели.",
     "После production PoP pipeline."),

    ("SECTION", "Эксплуатация и безопасность", "", "", ""),
    ("Emergency-управление", "⚪ Не начато",
     "Архитектура kill-switch (ADR-013): 4 уровня, fail-closed — доказана симулятором (41 тест).",
     "Нет emergency API endpoint'ов.\nНет MFA для emergency-пользователей.\nНет UI для отправки emergency-команд.\nНет мониторинга прогресса применения команд по устройствам.",
     "После device gateway stable."),
    ("Резервное копирование и DR", "⚪ Не начато",
     "—",
     "Нет автоматических бэкапов PostgreSQL.\nНет зеркалирования MinIO.\nНет DR-плана.\nНет нагрузочных тестов (ТЗ требует 40 000 устройств).",
     "Перед production deployment."),
    ("Мониторинг и алерты", "🟠 Частично",
     "Health endpoints: /health/live + /health/ready на всех сервисах.\nHealthState counters (published, failed, dead-letter, acked, nakd).\nGraceful shutdown (SIGTERM).",
     "Нет Prometheus metrics.\nНет alert rules.\nНет Grafana dashboards.\nНет мониторинга инфраструктуры (CPU, RAM, диск).",
     "Deferred: Prometheus + Grafana после compose → production."),
]

row = 2
for item in biz:
    if item[0] == "SECTION":
        for c in range(1, 6):
            ws2.cell(row=row, column=c, value=item[c-1])
        style_section_row(ws2, row, 5)
    else:
        write_cell(ws2, row, 1, item[0])
        write_status(ws2, row, 2, item[1])
        write_cell(ws2, row, 3, item[2])
        write_cell(ws2, row, 4, item[3])
        write_cell(ws2, row, 5, item[4])
    row += 1

# Freeze panes
ws1.freeze_panes = "A2"
ws2.freeze_panes = "A2"

# Save
output_path = "/home/cobalt/retail-media-platform-enterprise/docs/product/roadmap-s020-2026-07-10.xlsx"
import os
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Technical: {ws1.max_row} rows")
print(f"Business: {ws2.max_row} rows")
