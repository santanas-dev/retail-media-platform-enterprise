#!/usr/bin/env python3
"""
Roadmap-Consistency Guard — ROADMAP-GUARD-002.

Reads feature-registry.yaml, scans tests/ui-smoke/ for existing smoke tests,
and reads roadmap.xlsx (Бизнес-функции Roadmap) with its 4-column structure
(Бэкенд, UI, Юзер-стори (journey), Итог) to detect violations.

Two directions:
  A — Reachable features must not be understated in roadmap.
      If registry says reachable + smoke exists, roadmap must reflect it.
  B — Roadmap "Итог = Готово/Юзабельно" must be honest.
      Backend✅ + UI✅ + Story✅ = mandatory. No overclaim.

Modes:
  --audit   (default) Find violations, print findings, exit 0 (non-blocking).
  --strict  Exit 1 if any violation found (blocking CI gate).

Usage:
  python3 scripts/roadmap-consistency-check.py          # audit mode
  python3 scripts/roadmap-consistency-check.py --strict # blocking mode
"""

import argparse
import ast
import os
import re
import sys
from pathlib import Path

import openpyxl
import yaml

REPO_ROOT = Path(__file__).resolve().parent.parent
REGISTRY_PATH = REPO_ROOT / "docs" / "product" / "feature-registry.yaml"
ROADMAP_PATH = REPO_ROOT / "docs" / "product" / "roadmap-s020-2026-07-10.xlsx"
UI_SMOKE_DIR = REPO_ROOT / "tests" / "ui-smoke"

# Column names in the 4-column business sheet (ROADMAP-DONE-GATE-001)
COL_BACKEND = "Бэкенд"
COL_UI = "UI"
COL_STORY = "Юзер-стори (journey)"
COL_RESULT = "Итог"
COL_FUNC = "Бизнес-функция"

# ---- Helpers ---------------------------------------------------------------

def load_registry():
    with open(REGISTRY_PATH) as f:
        data = yaml.safe_load(f)
    return data.get("features", [])


def scan_smoke_functions():
    """Scan tests/ui-smoke/ for def test_uismoke__* functions.
    Returns dict: {test_function_name: file_path}."""
    smoke_funcs = {}
    if not UI_SMOKE_DIR.is_dir():
        return smoke_funcs
    for pyfile in UI_SMOKE_DIR.glob("*.py"):
        if pyfile.name.startswith("__"):
            continue
        try:
            tree = ast.parse(pyfile.read_text())
            for node in ast.walk(tree):
                if isinstance(node, ast.FunctionDef) and node.name.startswith("test_uismoke__"):
                    smoke_funcs[node.name] = str(pyfile.relative_to(REPO_ROOT))
        except SyntaxError:
            pass
    return smoke_funcs


def load_roadmap_business():
    """Read Бизнес-функции Roadmap sheet. Returns list of row dicts."""
    wb = openpyxl.load_workbook(ROADMAP_PATH, data_only=True)
    ws = wb["Бизнес-функции Roadmap"]
    headers = [str(c.value or "") for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        func = row[1]  # col B = Бизнес-функция
        if func is None or str(func).strip() == "":
            continue
        d = {}
        for i, h in enumerate(headers):
            d[h] = str(row[i]).strip() if row[i] is not None else ""
        rows.append(d)
    wb.close()
    return rows


# ---- Registry validation (unchanged from UI-TRUTH-001B) --------------------

REQUIRED_FIELDS = ["id", "frontend", "name", "route", "path", "smoke",
                   "priority", "roles", "status"]
VALID_STATUSES = {"reachable", "blocked"}


def validate_registry(features, smoke_funcs):
    findings = []
    seen_ids = set()
    for f in features:
        fid = f.get("id", "<missing>")
        if fid in seen_ids:
            findings.append(f"REGISTRY: duplicate id '{fid}'")
        seen_ids.add(fid)
        for field in REQUIRED_FIELDS:
            if field not in f:
                findings.append(f"REGISTRY: '{fid}' missing field '{field}'")
        status = f.get("status", "")
        if status not in VALID_STATUSES:
            findings.append(f"REGISTRY: '{fid}' invalid status '{status}'")
        frontend = f.get("frontend", "")
        smoke = f.get("smoke", "")
        if frontend != "service" and status == "reachable":
            if not smoke or smoke not in smoke_funcs:
                findings.append(
                    f"REGISTRY-SMOKE: '{fid}' status=reachable (UI) "
                    f"but smoke '{smoke}' not found in tests/ui-smoke/"
                )
    return findings


# ---- Roadmap → Registry consistency (4-column structure) -------------------

SERVICE_KEYWORDS = {
    "манифест", "manifest", "proof-of-play", "pop", "отчёт",
    "emergency", "резервн", "мониторинг", "observability",
    "kill-switch", "каналы", "устройств", "ксо", "android", "esl",
    "led", "price checker", "clickhouse", "billing", "экспорт",
    "dr /", "часовые пояс", "sla", "недопоказ", "операционн",
    "staged rollout", "feature flag", "data governance",
    "нагрузочн", "channel orchestrator", "siem", "интеграция с укм",
    "sales lift", "attribution", "self-service",
    "competitive", "store/audience", "financial docs",
    "programmatic", "dynamic creative", "mobile field",
    "independent dooh",
}

UI_SERVICE_MAP = {
    # Business function name → {feature_ids, combined: bool}
    "Вход сотрудников / рекламодателей": {
        "ids": ["self.login"],
    },
    "Роли и права (RBAC)": {
        "ids": ["user.assign_roles"],
    },
    "Личный кабинет рекламодателя": {
        "ids": ["self.login", "self.campaign_view", "self.report_view",
                "self.apply_or_brief", "self.campaign_create"],
    },
    "Создание и редактирование кампаний": {
        "ids": ["campaign.create", "campaign.edit", "campaign.submit",
                "campaign.activate", "campaign.pause", "campaign.complete"],
    },
    "Согласование кампаний (Approval)": {
        "ids": ["campaign.approve", "campaign.reject"],
    },
    "Загрузка креативов (медиафайлы)": {
        "ids": ["creative.upload", "creative.moderate_approve", "creative.moderate_reject"],
    },
    "Инвентарь": {
        "ids": ["inventory.simulate", "inventory.rule_create"],
    },
    "Управление рекламодателями": {
        "ids": ["advertiser.create_org", "advertiser.view",
                "advertiser.application_review", "advertiser.invite", "advertiser.apply"],
    },
}


def is_service_row(func_name):
    """Return True if this business row is a pure service/backend function."""
    name_lower = func_name.lower()
    return any(kw in name_lower for kw in SERVICE_KEYWORDS)


def parse_story_cell(story_raw):
    """Parse 'Юзер-стори (journey)' cell into {journey_id: status_char}.
    Examples: '✅ campaign.create / ⚪️ campaign.edit' →
              {'campaign.create': '✅', 'campaign.edit': '⚪️'}
              '✅ user.assign_roles' → {'user.assign_roles': '✅'}
    """
    result = {}
    if not story_raw or story_raw in ("—", "n/a", ""):
        return result
    # Split on / then extract ✅/⚪️/❌ + journey_id
    parts = re.split(r'\s*/\s*', story_raw)
    for part in parts:
        part = part.strip()
        m = re.match(r'([✅⚪️❌🟠])\s*([\w.]+)', part)
        if m:
            result[m.group(2)] = m.group(1)
    return result


def check_roadmap_vs_registry(roadmap_rows, features, smoke_funcs):
    findings = []

    # Build lookup
    feature_map = {f["id"]: f for f in features}
    reachable_ids = {f["id"] for f in features if f.get("status") == "reachable"}

    for row in roadmap_rows:
        func_name = row.get(COL_FUNC, "").strip()
        backend = row.get(COL_BACKEND, "").strip()
        ui = row.get(COL_UI, "").strip()
        story_raw = row.get(COL_STORY, "").strip()
        result = row.get(COL_RESULT, "").strip()

        if not func_name:
            continue

        # Skip service rows — they don't have UI smoke
        if is_service_row(func_name):
            continue

        # Skip rows not in UI_SERVICE_MAP (can't map to registry)
        map_entry = UI_SERVICE_MAP.get(func_name)
        if not map_entry:
            continue

        expected_ids = map_entry.get("ids", [])
        if not expected_ids:
            continue

        story_map = parse_story_cell(story_raw)

        # ── Direction A: Reachable features must not be understated ──
        for fid in expected_ids:
            feat = feature_map.get(fid)
            if not feat:
                continue
            if feat.get("status") != "reachable":
                continue
            if feat.get("frontend", "") == "service":
                continue

            smoke_name = feat.get("smoke", "")
            if not smoke_name or smoke_name not in smoke_funcs:
                continue  # registry validation catches this separately

            # Now: this feature is reachable + has green smoke.
            # Roadmap row must reflect it.
            story_status = story_map.get(fid)
            if not story_status or story_status != "✅":
                findings.append(
                    f"ROADMAP-UNDERSTATE: '{func_name}' — registry feature "
                    f"'{fid}' is reachable (green smoke), but story cell "
                    f"does not show ✅ (got: '{story_status or 'missing'}')"
                )

            # UI column should have ✅ (or mixed for combined rows)
            if "✅" not in ui:
                findings.append(
                    f"ROADMAP-UNDERSTATE: '{func_name}' — registry feature "
                    f"'{fid}' is reachable, but UI column='{ui}' "
                    f"(expected ✅ or mixed ✅/⚪️ for combined rows)"
                )

        # ── Direction B: Итог = Готово must be honest ──
        is_gotovo = result.startswith("✅ Готово") or result.startswith("✅ Готово/Юзабельно")
        if not is_gotovo:
            continue

        # Backend must be ✅
        if "✅" not in backend:
            findings.append(
                f"ROADMAP-OVERCLAIM: '{func_name}' — Итог='{result}' "
                f"but Бэкенд='{backend}' (must be ✅)"
            )

        # UI must be ✅
        if "✅" not in ui:
            findings.append(
                f"ROADMAP-OVERCLAIM: '{func_name}' — Итог='{result}' "
                f"but UI='{ui}' (must be ✅)"
            )

        # Every expected journey must be ✅ in story
        for fid in expected_ids:
            feat = feature_map.get(fid)
            if not feat:
                continue
            if feat.get("frontend", "") == "service":
                continue

            story_status = story_map.get(fid)
            if story_status != "✅":
                findings.append(
                    f"ROADMAP-OVERCLAIM: '{func_name}' — Итог='{result}' "
                    f"but journey '{fid}' has story status "
                    f"'{story_status or 'missing'}' (must be ✅)"
                )

            # Each ✅ journey must be reachable in registry
            if story_status == "✅" and fid not in reachable_ids:
                findings.append(
                    f"ROADMAP-OVERCLAIM: '{func_name}' — Итог='{result}' "
                    f"but journey '{fid}' marked ✅ in story while "
                    f"registry status={feat.get('status', '?')}"
                )

            # Each ✅ UI journey must have smoke
            smoke_name = feat.get("smoke", "")
            if story_status == "✅" and smoke_name and smoke_name not in smoke_funcs:
                findings.append(
                    f"ROADMAP-OVERCLAIM: '{func_name}' — Итог='{result}' "
                    f"but journey '{fid}' marked ✅ in story while "
                    f"smoke '{smoke_name}' not found"
                )

    return findings


# ---- Main ------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Roadmap-consistency guard (4-col)")
    parser.add_argument("--strict", action="store_true",
                        help="Exit 1 on any violation (blocking CI gate)")
    args = parser.parse_args()

    all_findings = []

    # Load data
    try:
        features = load_registry()
    except Exception as e:
        print(f"FATAL: cannot load registry: {e}", file=sys.stderr)
        sys.exit(2 if args.strict else 0)

    try:
        smoke_funcs = scan_smoke_functions()
    except Exception as e:
        print(f"WARNING: cannot scan smoke tests: {e}", file=sys.stderr)
        smoke_funcs = {}

    roadmap_rows = []
    roadmap_error = None
    try:
        roadmap_rows = load_roadmap_business()
    except Exception as e:
        roadmap_error = str(e)
        all_findings.append(f"ROADMAP-PARSE: cannot read roadmap.xlsx — {e}")

    # 1. Registry validation
    registry_findings = validate_registry(features, smoke_funcs)
    all_findings.extend(registry_findings)

    # 2. Roadmap vs registry (4-column)
    if not roadmap_error:
        roadmap_findings = check_roadmap_vs_registry(roadmap_rows, features, smoke_funcs)
        all_findings.extend(roadmap_findings)

    # Report
    print("=== Roadmap-Consistency Guard (ROADMAP-GUARD-002, 4-column) ===")
    print(f"  Registry: {len(features)} features")
    print(f"  Smoke tests found: {len(smoke_funcs)} functions")
    print(f"  Roadmap rows (business): {len(roadmap_rows)}")
    print(f"  Findings: {len(all_findings)}")
    print()

    if all_findings:
        print("--- Findings ---")
        for i, finding in enumerate(all_findings, 1):
            print(f"  [{i}] {finding}")
        print()
        print(f"SUMMARY: {len(all_findings)} violation(s) found.")
    else:
        print("SUMMARY: 0 violations — roadmap ↔ registry ↔ smoke consistent.")

    # Behavioral proof
    print()
    print("--- Behavioral Proof ---")
    reachable_ids = {f["id"] for f in features if f.get("status") == "reachable"}
    ui_reachable = [fid for fid in reachable_ids
                    if feature_map_safe(features, fid, "frontend") != "service"]
    print(f"  Reachable UI features: {len(ui_reachable)} — {ui_reachable}")
    for fid in ui_reachable:
        feat = feature_map_safe(features, fid, None)
        smoke = feat.get("smoke", "") if feat else ""
        in_smoke = smoke in smoke_funcs
        print(f"    {fid}: smoke={smoke} {'✅ found' if in_smoke else '❌ MISSING'}")
    print(f"  Reachable service features: "
          f"{len(reachable_ids) - len(ui_reachable)}")

    exit_code = 0
    if args.strict and all_findings:
        exit_code = 1
    sys.exit(exit_code)


def feature_map_safe(features, fid, key):
    for f in features:
        if f.get("id") == fid:
            return f.get(key) if key else f
    return None


if __name__ == "__main__":
    main()
