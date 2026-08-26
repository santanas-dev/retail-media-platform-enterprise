#!/usr/bin/env python3
"""Reproducible counters for the RM-GOV-002 migration.

Read-only. Recomputes the legacy-source counts directly from the XLSX and
cross-checks them against docs/product/roadmap-migration-manifest.yaml, so the
numbers in the manifest can never drift from the data they describe.

Usage:
    python3 scripts/ci/roadmap-migration-counts.py

Exit: 0 if counts agree, 1 on any mismatch.
"""

import sys
from collections import Counter
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parent.parent.parent
XLSX = ROOT / "docs" / "product" / "history" / "roadmap-s020-2026-07-10.xlsx"
MANIFEST = ROOT / "docs" / "product" / "roadmap-migration-manifest.yaml"


def source_counts():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    tech = wb["Технический Roadmap"]
    rows = [r for r in tech.iter_rows(min_row=2) if any(c.value for c in r)]
    section = [r for r in rows if str(r[0].value or "").strip().upper() == "SECTION"]
    biz = wb["Бизнес-функции Roadmap"]
    brows = [r for r in biz.iter_rows(min_row=2) if any(c.value for c in r)]
    return {
        "technical_items": len(rows) - len(section),
        "section_rows": len(section),
        "business_rows": len(brows),
    }


def main():
    src = source_counts()
    print("Из XLSX напрямую:")
    for k, v in src.items():
        print(f"    {k:18} {v}")

    if not MANIFEST.exists():
        print(f"\n[roadmap-migration-counts] FAIL — {MANIFEST.name} отсутствует")
        return 1

    man = yaml.safe_load(MANIFEST.read_text(encoding="utf-8"))
    declared = man.get("counts", {})
    entries = man.get("entries", [])

    actual = Counter()
    for e in entries:
        if e["source"] == "business-sheet":
            actual["business_rows"] += 1
        elif str(e.get("name", "")).upper() == "SECTION":
            actual["section_rows"] += 1
        else:
            actual["technical_items"] += 1

    print("\nВ манифесте (объявлено / фактически записей):")
    bad = False
    for k in src:
        d, a = declared.get(k), actual.get(k, 0)
        mark = "ok" if d == a == src[k] else "MISMATCH"
        if mark != "ok":
            bad = True
        print(f"    {k:18} declared={d} entries={a} source={src[k]}  {mark}")

    print("\nРаспределение решений:")
    for k, v in sorted(Counter(e["disposition"] for e in entries).items()):
        print(f"    {k:18} {v}")
    manual = sum(1 for e in entries if "[ручная поправка]" in e.get("reason", ""))
    open_out = sum(1 for e in entries if "ОТКРЫТЫЙ" in e.get("reason", ""))
    print(f"    {'ручных поправок':18} {manual}")
    print(f"    {'открытых вне очереди':18} {open_out}")

    if bad:
        print("\n[roadmap-migration-counts] FAIL — счётчики разошлись с источником")
        return 1
    print("\n[roadmap-migration-counts] PASS — счётчики совпадают с источником")
    return 0


if __name__ == "__main__":
    sys.exit(main())
