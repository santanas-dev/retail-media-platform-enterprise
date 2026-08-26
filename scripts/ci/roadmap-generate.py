#!/usr/bin/env python3
"""RM-GOV-003 — one-way roadmap projection generator.

SSOT inputs (never written by this script):
  docs/product/roadmap.yaml           — sequencing SSOT   (RM-GOV-001/002)
  docs/product/feature-registry.yaml  — functional SSOT   (reachable/blocked)
  tests/ui-smoke/*.py + ci-subset.txt — evidence          (what CI actually enforces)

Generated projections (read-only, regenerated wholesale):
  docs/product/generated/roadmap.generated.md
  docs/product/generated/roadmap.generated.xlsx
  docs/product/generated/roadmap-metrics.generated.json

These are the live roadmap views. The documents they replaced — `roadmap.md` and
`roadmap-s020-2026-07-10.xlsx` — were archived under `docs/product/history/` by the
RM-GOV-005 canonical cutover; this script asserts it never writes to them.

Determinism: no wall clock, no locale, no dict-order dependence. The workbook
zip is rewritten with fixed entry order and fixed timestamps so that two runs
produce byte-identical files.

Honesty: the generator projects only what an input states. Maturity levels
above what feature-registry proves are NOT derived — if roadmap.yaml carries no
`maturity` block, the projection says `не заявлено`, it does not guess.

Usage:
  python3 scripts/ci/roadmap-generate.py                  # write projections
  python3 scripts/ci/roadmap-generate.py --check-clean-diff   # exit 1 on drift
  python3 scripts/ci/roadmap-generate.py --self-test          # tamper matrix
"""

from __future__ import annotations

import argparse
import ast
import hashlib
import io
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

import yaml

REPO_ROOT = Path(__file__).resolve().parents[2]

ROADMAP_YAML = Path("docs/product/roadmap.yaml")
REGISTRY_YAML = Path("docs/product/feature-registry.yaml")
UI_SMOKE_DIR = Path("tests/ui-smoke")
CI_SUBSET = UI_SMOKE_DIR / "ci-subset.txt"
SMOKE_PREFIX = "test_uismoke__"

OUT_DIR = Path("docs/product/generated")
OUT_MD = OUT_DIR / "roadmap.generated.md"
OUT_XLSX = OUT_DIR / "roadmap.generated.xlsx"
OUT_METRICS = OUT_DIR / "roadmap-metrics.generated.json"

INPUTS = (ROADMAP_YAML, REGISTRY_YAML, CI_SUBSET)
OUTPUTS = (OUT_MD, OUT_XLSX, OUT_METRICS)

# Canon that this task must not touch before Gate G / RM-GOV-005.
PROTECTED = (
    Path("docs/product/history/roadmap-2026-08-26.md"),
    Path("docs/product/history/roadmap-s020-2026-07-10.xlsx"),
    Path("docs/product/feature-registry.yaml"),
    Path("PROJECT_STATE.md"),
    Path("AGENTS.md"),
)

# Fixed epoch for zip entries and document properties — derived from nothing
# that changes between runs.
FIXED_ZIP_DATE = (1980, 1, 1, 0, 0, 0)
FIXED_W3CDTF = "2026-08-26T00:00:00Z"

SHEET_TECH = "Технический Roadmap"
SHEET_BIZ = "Бизнес-функции Roadmap"

MATURITY_LADDER = [
    "implemented", "automated_verified", "ci_enforced", "stand_deployed",
    "stand_verified", "walkthrough_ok", "pilot_ready", "production_ready",
]
NOT_DECLARED = "не заявлено"


# --------------------------------------------------------------------------
# Input loading
# --------------------------------------------------------------------------

def load_inputs(root: Path) -> dict:
    roadmap = yaml.safe_load((root / ROADMAP_YAML).read_text(encoding="utf-8"))
    registry = yaml.safe_load((root / REGISTRY_YAML).read_text(encoding="utf-8"))
    features = registry.get("features", []) or []

    smoke_funcs = set()
    smoke_dir = root / UI_SMOKE_DIR
    if smoke_dir.is_dir():
        for pyfile in sorted(smoke_dir.glob("*.py")):
            if pyfile.name.startswith("__"):
                continue
            try:
                tree = ast.parse(pyfile.read_text(encoding="utf-8"))
            except SyntaxError:
                continue
            for node in ast.walk(tree):
                if isinstance(node, ast.FunctionDef) and node.name.startswith(SMOKE_PREFIX):
                    smoke_funcs.add(node.name)

    # ci-subset.txt stores SHORT names (campaign__create); feature-registry
    # stores full test function names (test_uismoke__campaign__create).
    # Both forms are indexed so the projection never depends on which one an
    # input happens to use.
    subset = set()
    subset_path = root / CI_SUBSET
    if subset_path.is_file():
        for line in subset_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            subset.add(line)
            subset.add(line if line.startswith(SMOKE_PREFIX) else SMOKE_PREFIX + line)

    return {
        "roadmap": roadmap,
        "features": features,
        "smoke_funcs": smoke_funcs,
        "ci_subset": subset,
    }


# --------------------------------------------------------------------------
# Derivation — every number below is computed, never transcribed
# --------------------------------------------------------------------------

def stage_order(roadmap: dict) -> dict:
    return {s["id"]: (s.get("order", 99), s["id"]) for s in roadmap.get("stages", [])}


def sorted_tasks(roadmap: dict) -> list:
    order = stage_order(roadmap)
    return sorted(roadmap.get("tasks", []),
                  key=lambda t: (order.get(t["stage"], (99, t["stage"])), t["id"]))


def dependency_depth(tasks: list) -> dict:
    by_id = {t["id"]: t for t in tasks}
    depth: dict = {}

    def walk(tid, seen):
        if tid in depth:
            return depth[tid]
        if tid in seen or tid not in by_id:
            return 0
        seen = seen | {tid}
        deps = [d for d in by_id[tid].get("dependencies", []) or [] if d in by_id]
        depth[tid] = 0 if not deps else 1 + max(walk(d, seen) for d in deps)
        return depth[tid]

    for t in tasks:
        walk(t["id"], frozenset())
    return depth


def maturity_index(roadmap: dict) -> dict:
    out = {}
    for entry in roadmap.get("maturity", []) or []:
        out[entry["feature_id"]] = entry
    return out


def count_by(items, key):
    out = {}
    for it in items:
        out[key(it)] = out.get(key(it), 0) + 1
    return dict(sorted(out.items()))


def build_metrics(data: dict) -> dict:
    roadmap = data["roadmap"]
    tasks = sorted_tasks(roadmap)
    features = data["features"]
    subset = data["ci_subset"]
    mat = maturity_index(roadmap)
    depth = dependency_depth(tasks)

    blocked = [f for f in features if f.get("status") == "blocked"]
    reachable = [f for f in features if f.get("status") == "reachable"]
    ui_features = [f for f in features if f.get("frontend") != "service"]
    ci_enforced = [f for f in ui_features if f.get("smoke") in subset]

    yaml_blocked = {b["feature_id"] for b in roadmap.get("blocked_features", []) or []}
    registry_blocked = {f["id"] for f in blocked}

    return {
        "schema_version": roadmap.get("schema_version"),
        "base_git_sha": roadmap.get("base", {}).get("git_sha"),
        "sequencing": {
            "tasks_total": len(tasks),
            "by_stage": count_by(tasks, lambda t: t["stage"]),
            "by_kind": count_by(tasks, lambda t: t["kind"]),
            "by_decision_status": count_by(tasks, lambda t: t["decision_status"]),
            "by_delivery_status": count_by(tasks, lambda t: t["delivery_status"]),
            "owner_gated": sum(1 for t in tasks if (t.get("owner_gate") or {}).get("required")),
            "with_verified_evidence": sum(
                1 for t in tasks
                if any(e.get("status") == "verified" for e in t.get("evidence_refs", []) or [])
            ),
            "max_dependency_depth": max(depth.values()) if depth else 0,
            "gates": [g["id"] for g in roadmap.get("gates", [])],
            "owner_decisions": len(roadmap.get("owner_decisions", []) or []),
        },
        "features": {
            "total": len(features),
            "reachable": len(reachable),
            "blocked": len(blocked),
            "by_frontend": count_by(features, lambda f: f.get("frontend", "?")),
            "ui_features": len(ui_features),
            "ci_enforced": len(ci_enforced),
            "blocked_with_empty_gap": sum(1 for f in blocked if not (f.get("gap") or "").strip()),
        },
        "maturity": {
            "declared": len(mat),
            "by_level": count_by(list(mat.values()), lambda m: m["level"]) if mat else {},
            "not_declared": len([f for f in features if f["id"] not in mat]),
            "reason": (
                "roadmap.yaml не содержит блока `maturity`; уровни выше registry "
                "не выводимы из входов и не домысливаются генератором"
            ) if not mat else "",
        },
        "cross_checks": {
            "blocked_in_yaml_not_in_registry": sorted(yaml_blocked - registry_blocked),
            "blocked_in_registry_not_in_yaml": sorted(registry_blocked - yaml_blocked),
            "smoke_declared_but_missing": sorted(
                f["id"] for f in ui_features
                if f.get("status") == "reachable" and f.get("smoke") not in data["smoke_funcs"]
            ),
        },
    }


# --------------------------------------------------------------------------
# Markdown projection
# --------------------------------------------------------------------------

def md_escape(text) -> str:
    return str(text if text is not None else "").replace("|", "\\|").replace("\n", " ").strip()


def fmt_acceptance(task: dict) -> str:
    parts = []
    for a in task.get("acceptance", []) or []:
        chunk = a["check"]
        vb = a.get("verified_by")
        if vb:
            chunk += f" [{vb}"
            if a.get("ref"):
                chunk += f": `{a['ref']}`"
            chunk += "]"
        parts.append(chunk)
    return "; ".join(parts)


def fmt_evidence(task: dict) -> str:
    refs = task.get("evidence_refs", []) or []
    if not refs:
        return "—"
    return "; ".join(f"{e['status']} · {e['kind']} · `{e['ref']}`" for e in refs)


def render_markdown(data: dict, metrics: dict) -> str:
    roadmap = data["roadmap"]
    tasks = sorted_tasks(roadmap)
    depth = dependency_depth(tasks)
    mat = maturity_index(roadmap)
    subset = data["ci_subset"]
    L = []

    L.append("# Roadmap — сгенерированная проекция")
    L.append("")
    L.append("> **СГЕНЕРИРОВАНО. НЕ РЕДАКТИРОВАТЬ РУКАМИ.**")
    L.append("> Правки вносятся в SSOT-входы, затем проекция перегенерируется.")
    L.append(">")
    L.append("> | | |")
    L.append("> |---|---|")
    L.append("> | Генератор | `scripts/ci/roadmap-generate.py` (RM-GOV-003) |")
    L.append(f"> | Входы | `{ROADMAP_YAML}`, `{REGISTRY_YAML}`, `{CI_SUBSET}` |")
    L.append(f"> | Base SHA | `{metrics['base_git_sha']}` |")
    L.append("> | Перегенерация | `python3 scripts/ci/roadmap-generate.py` |")
    L.append("> | Проверка дрейфа | `python3 scripts/ci/roadmap-generate.py --check-clean-diff` |")
    L.append("> | Статус | действующее представление roadmap (cutover `RM-GOV-005` выполнен) |")
    L.append("")
    L.append("Эти файлы — единственное действующее представление roadmap. "
             "Вытесненные `roadmap.md` и `roadmap-s020-2026-07-10.xlsx` архивированы "
             "в `docs/product/history/` (canonical cutover RM-GOV-005).")
    L.append("")

    # ---- metrics
    seq = metrics["sequencing"]
    feat = metrics["features"]
    L.append("## Метрики (посчитаны генератором)")
    L.append("")
    L.append("### Очередь")
    L.append("")
    L.append("| Метрика | Значение |")
    L.append("|---|---|")
    L.append(f"| Всего задач | {seq['tasks_total']} |")
    L.append(f"| По этапам | {', '.join(f'{k}={v}' for k, v in seq['by_stage'].items())} |")
    L.append(f"| По типу | {', '.join(f'{k}={v}' for k, v in seq['by_kind'].items())} |")
    L.append(f"| По статусу поставки | {', '.join(f'{k}={v}' for k, v in seq['by_delivery_status'].items())} |")
    L.append(f"| Требуют owner gate | {seq['owner_gated']} |")
    L.append(f"| С verified evidence | {seq['with_verified_evidence']} |")
    L.append(f"| Максимальная глубина зависимостей | {seq['max_dependency_depth']} |")
    L.append(f"| Гейты | {', '.join(seq['gates'])} |")
    L.append(f"| Решения владельца | {seq['owner_decisions']} |")
    L.append("")
    L.append("### Функции (из registry — функциональный SSOT)")
    L.append("")
    L.append("| Метрика | Значение |")
    L.append("|---|---|")
    L.append(f"| Всего функций | {feat['total']} |")
    L.append(f"| reachable · blocked | {feat['reachable']} · {feat['blocked']} |")
    L.append(f"| По фронтенду | {', '.join(f'{k}={v}' for k, v in feat['by_frontend'].items())} |")
    L.append(f"| UI-функций (не service) | {feat['ui_features']} |")
    L.append(f"| Закреплено в CI-subset | {feat['ci_enforced']} |")
    L.append(f"| blocked с пустым `gap` | {feat['blocked_with_empty_gap']} |")
    L.append("")
    m = metrics["maturity"]
    L.append("### Зрелость")
    L.append("")
    if m["declared"] == 0:
        L.append(f"Уровни зрелости **{NOT_DECLARED}** ни для одной функции.")
        L.append("")
        L.append(f"Причина: {m['reason']}.")
        L.append("")
        L.append("Генератор не выводит `stand_verified`, `walkthrough_ok`, `pilot_ready` и "
                 "`production_ready` из registry: registry доказывает достижимость, а не зрелость. "
                 "Пока владелец не заполнит блок `maturity` в `roadmap.yaml`, проекция говорит "
                 "«не заявлено», а не «0» — это разные утверждения.")
    else:
        L.append(f"Заявлено уровней: {m['declared']}; без уровня: {m['not_declared']}.")
        L.append("")
        L.append("| Уровень | Функций |")
        L.append("|---|---|")
        for lvl in MATURITY_LADDER:
            if lvl in m["by_level"]:
                L.append(f"| `{lvl}` | {m['by_level'][lvl]} |")
    L.append("")

    # ---- cross checks
    cc = metrics["cross_checks"]
    L.append("### Сверки между входами")
    L.append("")
    L.append("| Сверка | Результат |")
    L.append("|---|---|")
    L.append(f"| blocked в roadmap.yaml, но не в registry | {', '.join(cc['blocked_in_yaml_not_in_registry']) or '—'} |")
    L.append(f"| blocked в registry, но без `unblocked_by` | {', '.join(cc['blocked_in_registry_not_in_yaml']) or '—'} |")
    L.append(f"| reachable без найденного smoke | {', '.join(cc['smoke_declared_but_missing']) or '—'} |")
    L.append("")

    # ---- owner decisions
    L.append("## Решения владельца")
    L.append("")
    L.append("| ID | Статус | Дата | Формулировка |")
    L.append("|---|---|---|---|")
    for d in sorted(roadmap.get("owner_decisions", []) or [], key=lambda x: x["id"]):
        L.append(f"| `{d['id']}` | {d['status']} | {d.get('decided_on', '—')} | {md_escape(d['statement'])} |")
    L.append("")

    # ---- gates
    L.append("## Гейты")
    L.append("")
    L.append("| Гейт | Закрывает этап | Утверждает | Условия |")
    L.append("|---|---|---|---|")
    for g in sorted(roadmap.get("gates", []), key=lambda x: x["id"]):
        L.append(f"| `{g['id']}` | {g['closes_stage']} | {g['approver']} | "
                 f"{md_escape('; '.join(g.get('conditions', [])))} |")
    L.append("")

    # ---- queue per stage
    L.append("## Очередь по этапам")
    L.append("")
    stages = sorted(roadmap.get("stages", []), key=lambda s: (s.get("order", 99), s["id"]))
    for st in stages:
        st_tasks = [t for t in tasks if t["stage"] == st["id"]]
        if not st_tasks:
            continue
        closer = f" · закрывается `{st['closed_by_gate']}`" if st.get("closed_by_gate") else ""
        L.append(f"### {st['id']} — {st['title']} ({len(st_tasks)}){closer}")
        L.append("")
        L.append("| ID | Kind | Задача | Зависит от | Глубина | Поставка | Owner gate | Приёмка | Evidence |")
        L.append("|---|---|---|---|---|---|---|---|---|")
        for t in st_tasks:
            deps = ", ".join(f"`{d}`" for d in t.get("dependencies", []) or []) or "—"
            og = t.get("owner_gate") or {}
            og_txt = og.get("reason", "—") if og.get("required") else "—"
            alias = f" ({t['alias']})" if t.get("alias") else ""
            L.append(
                f"| `{t['id']}`{alias} | {t['kind']} | {md_escape(t['title'])} | {deps} | "
                f"{depth.get(t['id'], 0)} | {t['delivery_status']} | {og_txt} | "
                f"{md_escape(fmt_acceptance(t))} | {md_escape(fmt_evidence(t))} |"
            )
        L.append("")

    # ---- blocked features
    L.append("## Заблокированные функции и условия разблокировки")
    L.append("")
    L.append("| Feature | Registry gap | Разблокирует | Условия | Решение |")
    L.append("|---|---|---|---|---|")
    by_fid = {f["id"]: f for f in data["features"]}
    for b in sorted(roadmap.get("blocked_features", []) or [], key=lambda x: x["feature_id"]):
        f = by_fid.get(b["feature_id"], {})
        gap = (f.get("gap") or "").strip() or "**пусто в registry**"
        L.append(
            f"| `{b['feature_id']}` | {md_escape(gap)} | "
            f"{', '.join(f'`{u}`' for u in b['unblocked_by'])} | "
            f"{md_escape('; '.join(b['conditions']))} | {b.get('owner_decision', '—')} |"
        )
    L.append("")

    # ---- feature matrix
    L.append("## Матрица функций (registry + evidence)")
    L.append("")
    L.append("Столбец «Зрелость» берётся только из блока `maturity` в `roadmap.yaml`. "
             "Генератор его не вычисляет.")
    L.append("")
    for frontend in sorted({f.get("frontend", "?") for f in data["features"]}):
        group = sorted([f for f in data["features"] if f.get("frontend") == frontend],
                       key=lambda f: f["id"])
        L.append(f"### {frontend} — {len(group)}")
        L.append("")
        L.append("| Feature ID | Название | Приоритет | Статус | Smoke | В CI-subset | Зрелость |")
        L.append("|---|---|---|---|---|---|---|")
        for f in group:
            smoke = f.get("smoke") or "—"
            in_subset = "✅" if smoke in subset else "—"
            lvl = mat.get(f["id"], {}).get("level", NOT_DECLARED)
            L.append(
                f"| `{f['id']}` | {md_escape(f.get('name'))} | {f.get('priority', '—')} | "
                f"{f.get('status', '—')} | `{smoke}` | {in_subset} | {lvl} |"
            )
        L.append("")

    return "\n".join(L).rstrip() + "\n"


# --------------------------------------------------------------------------
# XLSX projection
# --------------------------------------------------------------------------

def render_xlsx(data: dict, metrics: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    roadmap = data["roadmap"]
    tasks = sorted_tasks(roadmap)
    depth = dependency_depth(tasks)
    mat = maturity_index(roadmap)
    subset = data["ci_subset"]

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    biz_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    wrap = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()

    def write_sheet(ws, headers, rows, widths, fill):
        ws.append(headers)
        for i, c in enumerate(ws[1], start=1):
            c.font = header_font
            c.fill = fill
            c.alignment = wrap
            ws.column_dimensions[c.column_letter].width = widths[i - 1]
        ws.freeze_panes = "A2"
        for r in rows:
            ws.append(r)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = wrap

    # -- Sheet 1: sequencing
    ws1 = wb.active
    ws1.title = SHEET_TECH
    rows1 = []
    for t in tasks:
        og = t.get("owner_gate") or {}
        rows1.append([
            t["id"],
            t["stage"],
            t["kind"],
            t["title"] + (f" ({t['alias']})" if t.get("alias") else ""),
            ", ".join(t.get("dependencies", []) or []) or "—",
            depth.get(t["id"], 0),
            t["decision_status"],
            t["delivery_status"],
            og.get("reason", "") if og.get("required") else "",
            fmt_acceptance(t),
            fmt_evidence(t).replace("`", ""),
            t.get("notes", ""),
        ])
    write_sheet(
        ws1,
        ["ID", "Этап", "Kind", "Задача", "Зависит от", "Глубина", "Decision",
         "Поставка", "Owner gate", "Приёмка", "Evidence", "Заметки"],
        rows1,
        [16, 8, 16, 46, 24, 9, 12, 14, 20, 60, 52, 60],
        header_fill,
    )

    # -- Sheet 2: features
    ws2 = wb.create_sheet(SHEET_BIZ)
    unblock = {b["feature_id"]: b for b in roadmap.get("blocked_features", []) or []}
    rows2 = []
    for f in sorted(data["features"], key=lambda x: (x.get("frontend", ""), x["id"])):
        smoke = f.get("smoke") or ""
        b = unblock.get(f["id"], {})
        rows2.append([
            f["id"],
            f.get("frontend", ""),
            f.get("name", ""),
            f.get("priority", ""),
            f.get("route", ""),
            smoke,
            "да" if smoke in subset else "нет",
            "да" if smoke in data["smoke_funcs"] else "нет",
            f.get("status", ""),
            mat.get(f["id"], {}).get("level", NOT_DECLARED),
            (f.get("gap") or "").strip(),
            ", ".join(b.get("unblocked_by", [])),
            "; ".join(b.get("conditions", [])),
        ])
    write_sheet(
        ws2,
        ["Feature ID", "Фронтенд", "Бизнес-функция", "Приоритет", "Маршрут", "Smoke",
         "В CI-subset", "Smoke существует", "Статус (registry)", "Зрелость",
         "Gap (registry)", "Разблокирует", "Условия разблокировки"],
        rows2,
        [26, 16, 40, 11, 14, 44, 13, 18, 18, 16, 42, 20, 52],
        biz_fill,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return normalize_xlsx(buf.getvalue())


def normalize_xlsx(raw: bytes) -> bytes:
    """Rewrite the workbook zip deterministically.

    openpyxl stamps the current time into docProps/core.xml and into every zip
    entry header. Both are replaced with fixed values so that repeated runs on
    identical inputs produce identical bytes.
    """
    src = zipfile.ZipFile(io.BytesIO(raw))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst:
        for name in sorted(src.namelist()):
            payload = src.read(name)
            if name.endswith(".xml") or name.endswith(".rels"):
                text = payload.decode("utf-8")
                text = re.sub(
                    r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}(\.\d+)?Z?",
                    FIXED_W3CDTF, text)
                payload = text.encode("utf-8")
            info = zipfile.ZipInfo(name, date_time=FIXED_ZIP_DATE)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            dst.writestr(info, payload)
    src.close()
    return out.getvalue()


# --------------------------------------------------------------------------
# Generation driver
# --------------------------------------------------------------------------

def generate(root: Path) -> dict:
    """Return {relative_path: bytes} for every projection. Writes nothing."""
    data = load_inputs(root)
    metrics = build_metrics(data)
    md = render_markdown(data, metrics)
    xlsx = render_xlsx(data, metrics)
    metrics_json = json.dumps(metrics, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    return {
        OUT_MD: md.encode("utf-8"),
        OUT_XLSX: xlsx,
        OUT_METRICS: metrics_json.encode("utf-8"),
    }


def write(root: Path, artifacts: dict) -> list:
    (root / OUT_DIR).mkdir(parents=True, exist_ok=True)
    changed = []
    for rel, payload in artifacts.items():
        target = root / rel
        if not target.exists() or target.read_bytes() != payload:
            target.write_bytes(payload)
            changed.append(str(rel))
    return changed


def assert_one_way(root: Path, before: dict) -> list:
    """Inputs and canon must be byte-identical after a generation run."""
    violations = []
    for rel, digest in before.items():
        path = root / rel
        now = hashlib.sha256(path.read_bytes()).hexdigest() if path.exists() else None
        if now != digest:
            violations.append(f"ONE-WAY: генератор изменил `{rel}`")
    return violations


def snapshot(root: Path) -> dict:
    out = {}
    for rel in list(INPUTS) + list(PROTECTED):
        path = root / rel
        if path.exists():
            out[rel] = hashlib.sha256(path.read_bytes()).hexdigest()
    return out


def check_clean_diff(root: Path) -> list:
    findings = []
    before = snapshot(root)
    artifacts = generate(root)
    findings.extend(assert_one_way(root, before))

    for rel, payload in artifacts.items():
        target = root / rel
        if not target.exists():
            findings.append(f"MISSING: `{rel}` отсутствует — запустите генератор")
            continue
        actual = target.read_bytes()
        if actual != payload:
            findings.append(
                f"DRIFT: `{rel}` отличается от проекции входов "
                f"(на диске sha256={hashlib.sha256(actual).hexdigest()[:12]}, "
                f"сгенерировано {hashlib.sha256(payload).hexdigest()[:12]})"
            )

    second = generate(root)
    for rel in artifacts:
        if artifacts[rel] != second[rel]:
            findings.append(f"NONDETERMINISM: два прогона дали разные байты для `{rel}`")

    return findings


# --------------------------------------------------------------------------
# Self-test — tamper matrix
# --------------------------------------------------------------------------

def _sandbox(root: Path, tmp: Path) -> Path:
    work = tmp / "repo"
    work.mkdir()
    for rel in [ROADMAP_YAML, REGISTRY_YAML, CI_SUBSET]:
        dst = work / rel
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(root / rel, dst)
    smoke_src = root / UI_SMOKE_DIR
    if smoke_src.is_dir():
        for f in smoke_src.glob("*.py"):
            shutil.copy2(f, work / UI_SMOKE_DIR / f.name)
    for rel in PROTECTED:
        src = root / rel
        if src.exists():
            dst = work / rel
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dst)
    write(work, generate(work))
    return work


def self_test(root: Path) -> int:
    cases = []

    def case(name, fn, expect_findings: bool):
        with tempfile.TemporaryDirectory() as td:
            work = _sandbox(root, Path(td))
            fn(work)
            findings = check_clean_diff(work)
            got = bool(findings)
            ok = got == expect_findings
            cases.append((name, ok, findings[:1]))

    def edit_yaml(work, transform):
        p = work / ROADMAP_YAML
        p.write_text(transform(p.read_text(encoding="utf-8")), encoding="utf-8")

    case("baseline: чистая генерация — зелено", lambda w: None, False)
    case("детерминизм: повторная генерация — зелено",
         lambda w: write(w, generate(w)), False)
    case("input drift: заголовок задачи изменён — красно",
         lambda w: edit_yaml(w, lambda s: s.replace(
             "title: Reconciliation/migration manifest",
             "title: Reconciliation/migration manifest (tampered)")), True)
    case("input drift: статус поставки изменён — красно",
         lambda w: edit_yaml(w, lambda s: s.replace(
             "  delivery_status: planned", "  delivery_status: done", 1)), True)
    case("input drift: зависимость удалена — красно",
         lambda w: edit_yaml(w, lambda s: s.replace(
             "  - RM-GOV-002\n", "", 1)), True)
    case("registry drift: blocked → reachable — красно",
         lambda w: (w / REGISTRY_YAML).write_text(
             (w / REGISTRY_YAML).read_text(encoding="utf-8").replace(
                 "status: blocked", "status: reachable", 1), encoding="utf-8"), True)
    case("evidence drift: строка удалена из ci-subset — красно",
         lambda w: (w / CI_SUBSET).write_text("\n".join(
             [l for l in (w / CI_SUBSET).read_text(encoding="utf-8").splitlines()
              if l.strip().startswith("#") or not l.strip()][:5] +
             [l for l in (w / CI_SUBSET).read_text(encoding="utf-8").splitlines()
              if l.strip() and not l.strip().startswith("#")][1:]) + "\n",
             encoding="utf-8"), True)
    case("output tamper: правка руками в Markdown — красно",
         lambda w: (w / OUT_MD).write_text(
             (w / OUT_MD).read_text(encoding="utf-8").replace(
                 "| Всего задач | 42 |", "| Всего задач | 41 |"), encoding="utf-8"), True)
    case("output tamper: правка руками в metrics.json — красно",
         lambda w: (w / OUT_METRICS).write_text(
             (w / OUT_METRICS).read_text(encoding="utf-8").replace(
                 '"blocked": 5', '"blocked": 0'), encoding="utf-8"), True)
    case("output tamper: правка руками в XLSX — красно",
         lambda w: (w / OUT_XLSX).write_bytes(
             (w / OUT_XLSX).read_bytes() + b"\x00"), True)
    case("output tamper: проекция удалена — красно",
         lambda w: (w / OUT_MD).unlink(), True)

    # Non-invention guard: no maturity block in input → no ladder level in output.
    with tempfile.TemporaryDirectory() as td:
        work = _sandbox(root, Path(td))
        roadmap = yaml.safe_load((work / ROADMAP_YAML).read_text(encoding="utf-8"))
        has_maturity = bool(roadmap.get("maturity"))
        md_text = (work / OUT_MD).read_text(encoding="utf-8")
        invented = [] if has_maturity else [
            lvl for lvl in MATURITY_LADDER if f"| {lvl} |" in md_text
        ]
        ok = (not invented) and (has_maturity or NOT_DECLARED in md_text)
        cases.append(("не-выдумывание: без блока maturity уровни не проставлены", ok,
                      [f"проекция содержит уровни {invented}"] if invented else []))

    # One-way guard: canon untouched by a real run in the sandbox.
    with tempfile.TemporaryDirectory() as td:
        work = _sandbox(root, Path(td))
        before = snapshot(work)
        write(work, generate(work))
        viol = assert_one_way(work, before)
        cases.append(("one-way: канон и входы не изменены прогоном", not viol, viol[:1]))

    width = max(len(n) for n, _, _ in cases)
    failed = 0
    for name, ok, detail in cases:
        mark = "PASS" if ok else "FAIL"
        if not ok:
            failed += 1
        line = f"  [{mark}] {name.ljust(width)}"
        if not ok and detail:
            line += f"   ← {detail[0]}"
        print(line)
    print(f"\n[roadmap-generate] self-test: {len(cases) - failed}/{len(cases)} passed")
    return 1 if failed else 0


# --------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description="RM-GOV-003 roadmap projection generator")
    ap.add_argument("--root", default=str(REPO_ROOT), help="repository root")
    ap.add_argument("--check-clean-diff", action="store_true",
                    help="regenerate in memory and fail on any drift; writes nothing")
    ap.add_argument("--self-test", action="store_true", help="run the tamper matrix")
    args = ap.parse_args()
    root = Path(args.root).resolve()

    if args.self_test:
        return self_test(root)

    if args.check_clean_diff:
        findings = check_clean_diff(root)
        if findings:
            print(f"[roadmap-generate] FAIL — {len(findings)} расхождений:")
            for f in findings:
                print(f"  - {f}")
            return 1
        print("[roadmap-generate] CLEAN — проекции совпадают с входами, генерация детерминирована")
        return 0

    before = snapshot(root)
    artifacts = generate(root)
    changed = write(root, artifacts)
    viol = assert_one_way(root, before)
    if viol:
        for v in viol:
            print(f"[roadmap-generate] {v}")
        return 1
    if changed:
        print("[roadmap-generate] обновлено:")
        for c in changed:
            print(f"  - {c}")
    else:
        print("[roadmap-generate] без изменений — проекции уже актуальны")
    return 0


if __name__ == "__main__":
    sys.exit(main())
