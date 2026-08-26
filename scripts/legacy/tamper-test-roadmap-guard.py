#!/usr/bin/env python3
"""QUARANTINED — НЕ ЗАПУСКАТЬ.

Помещён в карантин canonical cutover RM-GOV-005 (2026-08-26).

Тест подменял строки в каноническом `docs/product/roadmap-s020-2026-07-10.xlsx`
и проверял направления registry ↔ книга. После cutover книга ГЕНЕРИРУЕТСЯ из
registry, поэтому эти направления не могут разойтись по построению, а сама книга
архивирована в `docs/product/history/`.

Живые проверки, которые здесь были, никуда не делись:
  orphan smoke, поле/статус registry, членство в ci-subset
    -> модуль `registry` гейта scripts/ci/roadmap-governance-guard.py
  подмена входа и ручная правка представления
    -> измерения drift и SSOT того же гейта, tamper-матрица --self-test

Отдельно ради истории: направление OVERCLAIM этого теста было красным на develop,
потому что его фикстура перестала быть ложью (RM-GOV-004 перенацелил её на строку,
которая ещё blocked). Урок закреплён в новом гейте проверкой инертности фикстур.
"""

"""Tamper tests for ROADMAP-GUARD-002.

Creates temporary copies of roadmap.xlsx with deliberate violations,
runs the guard, and asserts violations are detected.

Tests:
  1. UNDERSTATE: change G1 story from ✅ → ⚪️ → detected
  2. OVERCLAIM: set Итог=✅ on blocked row → detected
  3. CLEAN: unmodified → 0 findings
"""

import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
ROADMAP_ORIG = REPO_ROOT / "docs" / "product" / "roadmap-s020-2026-07-10.xlsx"
GUARD_SCRIPT = REPO_ROOT / "scripts" / "roadmap-consistency-check.py"
REGISTRY_ORIG = REPO_ROOT / "docs" / "product" / "feature-registry.yaml"

passed = 0
failed = 0


def run_guard_on(tamper_path):
    """Copy tampered file over real, run guard, restore."""
    backup = str(ROADMAP_ORIG) + ".bak"
    shutil.copy2(ROADMAP_ORIG, backup)
    try:
        shutil.copy2(tamper_path, ROADMAP_ORIG)
        result = subprocess.run(
            [sys.executable, str(GUARD_SCRIPT)],
            capture_output=True, text=True, timeout=30,
        )
        return result.stdout
    finally:
        shutil.copy2(backup, ROADMAP_ORIG)
        os.remove(backup)


def test(name, tamper_fn, expect_violation=True):
    global passed, failed
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        shutil.copy2(ROADMAP_ORIG, tmp.name)
        tp = tmp.name

    try:
        wb = openpyxl.load_workbook(tp)
        tamper_fn(wb)
        wb.save(tp)
        stdout = run_guard_on(tp)
        clean = "0 violations" in stdout

        if expect_violation and not clean:
            print(f"  ✅ PASS: {name}")
            passed += 1
        elif not expect_violation and clean:
            print(f"  ✅ PASS: {name} (clean)")
            passed += 1
        else:
            print(f"  ❌ FAIL: {name}")
            print(f"     Expected: {'violation' if expect_violation else 'clean'}")
            print(f"     Got: {'clean' if clean else 'violations found'}")
            failed += 1
    finally:
        try:
            os.unlink(tp)
        except OSError:
            pass


ws_name = "Бизнес-функции Roadmap"

print("=== ROADMAP-GUARD-002 Tamper Tests ===\n")


# Test 1: Understate G1
def understate_g1(wb):
    ws = wb[ws_name]
    for row in range(2, ws.max_row + 1):
        func = str(ws.cell(row=row, column=2).value or "")
        if "Создание и редактирование" in func:
            old = str(ws.cell(row=row, column=5).value or "")
            ws.cell(row=row, column=5).value = old.replace(
                "✅ campaign.create", "⚪️ campaign.create"
            )
            return

test("Understate G1 (campaign.create ✅→⚪️)", understate_g1)


# Test 2: Overclaim blocked row
#
# RM-GOV-004: the row this case used to target ("Согласование кампаний") stopped
# being an overclaim once campaign.approve/campaign.reject became reachable with
# green CI smokes — the tamper was no longer a lie, the guard correctly answered
# "clean", and this case had been silently proving nothing. Retargeted to a row
# that is still genuinely blocked in feature-registry.yaml, and made to assert
# that at the moment it runs, so it cannot go stale in silence again.
OVERCLAIM_ROW = "Личный кабинет рекламодателя"
OVERCLAIM_BLOCKED_IDS = ("self.report_view", "self.campaign_create")


def _assert_still_blocked():
    """Fail loudly if the chosen row stopped being an overclaim."""
    import yaml as _yaml
    reg = _yaml.safe_load(REGISTRY_ORIG.read_text())["features"]
    status = {f["id"]: f.get("status") for f in reg}
    live = [i for i in OVERCLAIM_BLOCKED_IDS if status.get(i) == "blocked"]
    if not live:
        print(f"  ❌ FAIL: фикстура устарела — ни одна из {OVERCLAIM_BLOCKED_IDS} "
              f"больше не blocked, подмена перестала быть ложью")
        return False
    return True


def overclaim_blocked(wb):
    ws = wb[ws_name]
    for row in range(2, ws.max_row + 1):
        func = str(ws.cell(row=row, column=2).value or "")
        if OVERCLAIM_ROW in func:
            ws.cell(row=row, column=6).value = "✅ Готово/Юзабельно"
            ws.cell(row=row, column=3).value = "✅"
            ws.cell(row=row, column=4).value = "✅"
            ws.cell(row=row, column=5).value = (
                "✅ self.login / ✅ self.campaign_view / ✅ self.report_view "
                "/ ✅ self.apply_or_brief / ✅ self.campaign_create"
            )
            return
    raise AssertionError(f"фикстура устарела: строка {OVERCLAIM_ROW!r} не найдена")


if _assert_still_blocked():
    test("Overclaim (кабинет рекламодателя Итог=✅ при blocked-функциях)", overclaim_blocked)
else:
    failed += 1


# Test 3: Clean
test("Clean workbook → 0 findings", lambda wb: None, expect_violation=False)

# Test 4 (Direction C): Orphan smoke — remove smoke ref from registry


def run_guard_direct():
    """Run guard directly (no tamper) — for registry tamper tests."""
    result = subprocess.run(
        [sys.executable, str(GUARD_SCRIPT)],
        capture_output=True, text=True, timeout=30,
    )
    return result.stdout


def test_registry_tamper(name, tamper_fn, expect_violation=True):
    """Tamper registry, run guard, restore."""
    global passed, failed
    backup = str(REGISTRY_ORIG) + ".bak"
    shutil.copy2(REGISTRY_ORIG, backup)
    try:
        tamper_fn()
        stdout = run_guard_direct()
        clean = "0 violations" in stdout

        if expect_violation and not clean:
            print(f"  ✅ PASS: {name}")
            passed += 1
        elif not expect_violation and clean:
            print(f"  ✅ PASS: {name} (clean)")
            passed += 1
        else:
            print(f"  ❌ FAIL: {name}")
            print(f"     Expected: {'violation' if expect_violation else 'clean'}")
            print(f"     Got: {'clean' if clean else 'violations found'}")
            failed += 1
    finally:
        shutil.copy2(backup, REGISTRY_ORIG)
        os.remove(backup)


def orphan_brand_smoke():
    """Remove smoke reference from advertiser.brand_crud."""
    with open(REGISTRY_ORIG) as f:
        content = f.read()
    # Change the smoke field to a non-existent name
    tampered = content.replace(
        "smoke: test_uismoke__advertiser__brand_crud",
        "smoke: test_uismoke__advertiser__brand_crud_NONEXISTENT"
    )
    if tampered == content:
        raise RuntimeError("Failed to find smoke reference to tamper")
    with open(REGISTRY_ORIG, "w") as f:
        f.write(tampered)


test_registry_tamper(
    "Direction C: orphan smoke → SMOKE-ORPHAN detected",
    orphan_brand_smoke,
    expect_violation=True,
)

print(f"\n=== Results: {passed} passed, {failed} failed ===")
sys.exit(0 if failed == 0 else 1)
