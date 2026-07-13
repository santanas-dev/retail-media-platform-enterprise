#!/usr/bin/env python3
"""S-020 QA fix: update roadmap with corrected admin-web status and other accuracy fixes."""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = load_workbook("/home/cobalt/retail-media-platform-enterprise/docs/product/roadmap-s020-2026-07-10.xlsx")

# Colors
done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
done_font = Font(color="006100", bold=True)
pilot_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
pilot_font = Font(color="9C6500", bold=True)
partial_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
partial_font = Font(color="974706", bold=True)
not_started_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
not_started_font = Font(color="808080")
deferred_fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
deferred_font = Font(color="5B3A8C")

thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)
wrap_align = Alignment(wrap_text=True, vertical='top')

def set_status(ws, row, col, status):
    cell = ws.cell(row=row, column=col)
    if status == "✅ Готово":
        cell.fill = done_fill; cell.font = done_font
    elif status == "🟡 Pilot-ready":
        cell.fill = pilot_fill; cell.font = pilot_font
    elif status == "🟠 Частично":
        cell.fill = partial_fill; cell.font = partial_font
    elif status == "🚫 Deferred":
        cell.fill = deferred_fill; cell.font = deferred_font
    else:
        cell.fill = not_started_fill; cell.font = not_started_font
    cell.value = status
    cell.alignment = wrap_align
    cell.border = thin_border

def set_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.alignment = wrap_align
    cell.border = thin_border

# =========================================================================
# Fix 1: Технический Roadmap — "Admin Campaign UI"
# =========================================================================
ws1 = wb["Технический Roadmap"]
for row in range(2, ws1.max_row + 1):
    name = ws1.cell(row=row, column=1).value
    if name and "Admin Campaign UI" in str(name):
        # Update status from 🟠 Частично to 🟡 Pilot-ready
        set_status(ws1, row, 2, "🟡 Pilot-ready")
        # Update evidence
        set_cell(ws1, row, 3,
            "Campaign list (real API with status filter). Create wizard "
            "(full form, reference pickers). Detail page with 5 tabs: "
            "flights, placements, creatives+upload, approvals, "
            "PoP reporting (summary/by-day/by-surface). "
            "64 vitest tests (5 files). 1473-line CampaignDetailPage.")
        # Update next step
        set_cell(ws1, row, 4, "Advertisers page (placeholder → real). Polish/UX/accessibility. Browser E2E.")
        # Update notes
        set_cell(ws1, row, 5, "Admin-web: 14 .tsx files, 64 tests. Auth + campaign CRUD UI + approval UI + PoP UI + creative upload UI all work.")

    # Fix "Admin Web Auth Shell" — update evidence
    if name and "Admin Web Auth Shell" in str(name):
        # Keep status 🟠 Частично for this row (auth is one piece of frontend)
        set_cell(ws1, row, 3,
            "React 19, react-router-dom, vitest. Login/logout/session, "
            "protected routes, /me from API, sidebar nav. "
            "Auth contract: HttpOnly cookie, no refresh_token in JSON body. "
            "17 tests (13 API contract + 4 auth-shell). Build: 48 modules, 605ms.")
        set_cell(ws1, row, 4, "— (auth shell complete)")

    # Fix "Campaign CRUD" — add UI note
    if name and "Campaign CRUD" in str(name):
        set_cell(ws1, row, 5,
            "Tenant isolation, cross-org validation, outbox. "
            "UI: CampaignListPage + CampaignCreatePage + CampaignDetailPage connected to real API.")

# =========================================================================
# Fix 2: Бизнес-функции — update "Создание и редактирование кампаний"
# =========================================================================
ws2 = wb["Бизнес-функции Roadmap"]
for row in range(2, ws2.max_row + 1):
    name = ws2.cell(row=row, column=1).value
    if name and "Создание и редактирование" in str(name):
        # Status stays 🟡 Pilot-ready — this is correct
        set_cell(ws2, row, 3,
            "Backend API: полный CRUD кампаний (7 endpoints).\n"
            "Admin-web UI: список кампаний с фильтром по статусу, "
            "мастер создания (выбор рекламодателя/бренда/контракта, "
            "часовой пояс), страница кампании с 5 вкладками:\n"
            "• Пролёты — создание и редактирование\n"
            "• Размещения — привязка к display surfaces\n"
            "• Креативы — привязка + загрузка файлов\n"
            "• Согласование — request/approve/reject\n"
            "• PoP-отчёты — summary, по дням, по поверхностям\n"
            "64 vitest теста.")
        set_cell(ws2, row, 4,
            "Нет страницы рекламодателей (заглушка).\n"
            "Нет календаря планирования.\n"
            "Нет drag-and-drop.\n"
            "UX/accessibility не завершены.\n"
            "Нет браузерных E2E-тестов.")

    # Fix "Согласование кампаний (Approval)" — UI done now
    if name and "Согласование" in str(name):
        set_cell(ws2, row, 3,
            "Backend: request-approval → approve/reject. Полный audit trail.\n"
            "Admin-web UI: вкладка «Согласование» на странице кампании — "
            "кнопки «Отправить», «Утвердить», «Отклонить» с модальным окном."
            "Лог статусов с датами и причинами.")

    # Fix "Загрузка креативов" — clarify not production-ready
    if name and "Загрузка креативов" in str(name):
        set_cell(ws2, row, 3,
            "Backend: presigned URL (MinIO/S3) с серверной проверкой SHA-256.\n"
            "Admin-web UI: кнопка «Загрузить файл», прогресс-бар.\n"
            "Безопасность: storage_bucket/key не экспозятся.\n"
            "Пилотное авто-одобрение (CREATIVE_AUTO_APPROVE_UPLOADS=true).")
        set_cell(ws2, row, 4,
            "Нет проверки на вирусы (malware scan).\n"
            "Нет ручной модерации креативов.\n"
            "Нет транскодирования и CDN.\n"
            "Файлы загружаются только через admin-web "
            "(нет рекламодательского портала).")
        set_cell(ws2, row, 5,
            "Malware scan, transcoding, manual moderation — отложены.\n"
            "Работает для пилота.")

    # Fix "Отчёты по показам" — add UI note
    if name and "Отчёты по показам" in str(name) and "PoP" not in str(name):
        set_cell(ws2, row, 3,
            "Backend API: GET pop/{summary,by-day,by-surface}.\n"
            "Admin-web UI: вкладка «PoP-отчёт» на странице кампании — "
            "сводка, график по дням, таблица по поверхностям.\n"
            "Scoped: рекламодатель видит только свои кампании.")
        set_cell(ws2, row, 5,
            "После campaign UI стабилизации.")

    # Fix "Вход сотрудников" — more precise
    if name and "Вход сотрудников" in str(name):
        set_cell(ws2, row, 3,
            "Тестовый вход через local_advertiser и local_break_glass (bcrypt).\n"
            "Admin-web: страница логина с выбором провайдера "
            "(Рекламодатель / Break-glass Admin / Сотрудник AD).\n"
            "JWT access + refresh токены, HttpOnly cookie. /me из БД.")

# =========================================================================
# Fix 3: "Device Gateway / manifest" — clarify pilot-ready, not production
# (Business sheet)
# =========================================================================
for row in range(2, ws2.max_row + 1):
    name = ws2.cell(row=row, column=1).value
    if name and "Получение manifest" in str(name):
        set_cell(ws2, row, 4,
            "Только HTTP (не HTTPS/mTLS).\n"
            "Подпись manifest — HMAC-заглушка.\n"
            "Нет emergency-канала для срочных команд.\n"
            "Нет реального KSO плеера, принимающего manifest.")
        set_cell(ws2, row, 5,
            "mTLS после интеграции с KSO плеером.\n"
            "Production manifest signing.")

    if name and "Формирование плейлистов" in str(name):
        set_cell(ws2, row, 4,
            "Подпись manifest — HMAC-заглушка (не для production).\n"
            "Nested per-surface playlist — Manifest v2.\n"
            "Манифест доставляется только в тестах (нет реального плеера).")

# Save
wb.save("/home/cobalt/retail-media-platform-enterprise/docs/product/roadmap-s020-2026-07-10.xlsx")

# Verify
wb2 = load_workbook("/home/cobalt/retail-media-platform-enterprise/docs/product/roadmap-s020-2026-07-10.xlsx")
for s in wb2.sheetnames:
    ws = wb2[s]
    statuses = {}; sections = 0
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        n, st = row
        if n and st:
            if n == 'SECTION': sections += 1
            else: statuses[st] = statuses.get(st, 0) + 1
    print(f"--- {s} ---")
    for st, c in sorted(statuses.items()):
        print(f"  {st}: {c}")

import os
print(f"\nSize: {os.path.getsize('/home/cobalt/retail-media-platform-enterprise/docs/product/roadmap-s020-2026-07-10.xlsx'):,} bytes")
print("QA fixes applied.")
