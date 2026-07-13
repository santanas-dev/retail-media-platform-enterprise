#!/usr/bin/env python3
"""Update roadmap xlsx with v2.6 next-branch rows."""
import openpyxl

wb = openpyxl.load_workbook('docs/product/roadmap-s020-2026-07-10.xlsx')

# -- Tech roadmap --
ws = wb['Технический Roadmap']
start_row = ws.max_row + 1

v26_tech = [
    ("SECTION", "v2.6 Next Branch", "", "", ""),
    ("Tenant Model ADR (ADR-018)", "\U0001f7e1 Decision needed",
     "ADR-018 proposed 2026-07-11. P0 before v2.6 implementation.",
     "Single-retailer vs multi-retailer decision.",
     "Without decision: rewrite attribution, finance, competitive separation, RLS."),
    ("Attribution & Sales Lift", "\u26aa Not started",
     "TZ v2.6. Future v2.6.", "After tenant model ADR.",
     "Requires PoP pipeline + receipt data."),
    ("Self-Service Advertiser Cabinet", "\U0001f7e0 Foundation only",
     "S-023 design gate: backend API ready, tenant isolation proven. apps/advertiser-web planned.",
     "Seed fix: advertiser role, scaffold, campaign list.",
     "Not v2.6 feature. This is TZ v2.5 portal."),
    ("Competitive Separation", "\u26aa Not started",
     "TZ v2.6. Future v2.6.", "After tenant model ADR.",
     "Blocking competitor brands on same screen."),
    ("Store-Level Audience Targeting", "\u26aa Not started",
     "TZ v2.6. Future v2.6.", "After tenant model ADR + inventory.",
     "Campaign-to-store-profile targeting."),
    ("Finance Contract/Invoicing Integration", "\u26aa Not started",
     "TZ v2.6. Future v2.6.", "After tenant model ADR.",
     "1C/ERP API integration."),
    ("Programmatic Extension Point", "\U0001f6ab Deferred",
     "TZ v2.6. OpenRTB stub.", "After production pilot.",
     "External DSP/SSP."),
    ("Dynamic Creative MVP", "\U0001f6ab Deferred",
     "TZ v2.6. HTML5 canvas/video templating.", "After KSO player stable.",
     "Creative personalization."),
    ("Mobile Field Ops MVP", "\U0001f6ab Deferred",
     "TZ v2.6. Mobile app for field operators.", "After KSO player stable.",
     "Screen checks, content swap in field."),
    ("A/B Lift Metrics", "\u26aa Not started",
     "TZ v2.6. Future v2.6.", "After attribution.",
     "A/B testing of creatives."),
    ("Third-Party DOOH Measurement/Accreditation", "\U0001f6ab Deferred",
     "TZ v2.6. Design stub only.", "After production pilot.",
     "Independent impression audit."),
]

for i, (a, b, c, d, e) in enumerate(v26_tech):
    row = start_row + i
    ws.cell(row=row, column=1, value=a)
    ws.cell(row=row, column=2, value=b)
    ws.cell(row=row, column=3, value=c)
    ws.cell(row=row, column=4, value=d)
    ws.cell(row=row, column=5, value=e)

print(f"Tech roadmap: added {len(v26_tech)} rows at row {start_row}")

# -- Biz roadmap --
ws2 = wb['Бизнес-функции Roadmap']
start_row2 = ws2.max_row + 1

v26_biz = [
    ("SECTION", "v2.6 Next Branch", "", "", ""),
    ("Self-service advertiser cabinet", "\u26aa Not started",
     "Foundation exists: backend API ready, tenant isolation proven (S-023 design gate). apps/advertiser-web planned.",
     "No separate portal. Admin-web has no role-based routing. No self-service campaign creation.",
     "S-023a: scaffold advertiser-web + seed fix."),
    ("Sales Lift / Attribution proof", "\u26aa Not started",
     "TZ v2.6. Future v2.6.",
     "No attribution pipeline. No receipt data integration.",
     "After tenant model ADR + PoP pipeline."),
    ("Competitor blocking", "\u26aa Not started",
     "TZ v2.6. Future v2.6.",
     "No competitive separation model.",
     "After tenant model ADR."),
    ("Store/audience targeting", "\u26aa Not started",
     "TZ v2.6. Future v2.6.",
     "No store profiles. No audience targeting.",
     "After tenant model ADR + inventory."),
    ("Financial docs / reconciliation / integration", "\u26aa Not started",
     "TZ v2.6. Future v2.6.",
     "No 1C/ERP API. No invoicing.",
     "After tenant model ADR."),
    ("Programmatic inventory sales", "\U0001f6ab Deferred",
     "TZ v2.6. OpenRTB stub.",
     "No DSP/SSP integration.",
     "After production pilot."),
    ("Dynamic creatives", "\U0001f6ab Deferred",
     "TZ v2.6. HTML5 canvas/video templating.",
     "No creative templating.",
     "After KSO player stable."),
    ("Mobile field ops", "\U0001f6ab Deferred",
     "TZ v2.6. Mobile app.",
     "No mobile client.",
     "After KSO player stable."),
    ("Independent DOOH measurement", "\U0001f6ab Deferred",
     "TZ v2.6. Design stub only.",
     "No external impression audit.",
     "After production pilot."),
]

for i, (a, b, c, d, e) in enumerate(v26_biz):
    row = start_row2 + i
    ws2.cell(row=row, column=1, value=a)
    ws2.cell(row=row, column=2, value=b)
    ws2.cell(row=row, column=3, value=c)
    ws2.cell(row=row, column=4, value=d)
    ws2.cell(row=row, column=5, value=e)

print(f"Biz roadmap: added {len(v26_biz)} rows at row {start_row2}")

# Save
wb.save('docs/product/roadmap-s020-2026-07-10.xlsx')
print("Saved.")

# Verify
wb2 = openpyxl.load_workbook('docs/product/roadmap-s020-2026-07-10.xlsx')
print(f"Verify: sheets={wb2.sheetnames}")
print(f"Tech max_row={wb2['Технический Roadmap'].max_row}")
print(f"Biz max_row={wb2['Бизнес-функции Roadmap'].max_row}")
